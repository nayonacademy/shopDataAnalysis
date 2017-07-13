from openpyxl import Workbook
import xml.etree.ElementTree as ET
tree = ET.parse('shop.xml')
root = tree.getroot()
wb = Workbook()
ws = wb.active

for idx, child in enumerate(root.findall('Product')):
    DistinctiveTitle = child.find('DistinctiveTitle').text
    idx = idx +2
    col = 'A'+ str(idx)
    print("col :: ", col)
    ws[col] = DistinctiveTitle

    for id in child.findall('ProductIdentifier'):
        ProductIDType = id.find('ProductIDType').text
        if ProductIDType == '15':
            IDValue = id.find('IDValue').text
            ws['B'+ str(idx)] = IDValue
            print("ProductIDType :: ", ProductIDType, IDValue)

    SupplyDetail = child.find('SupplyDetail')
    for price in SupplyDetail.findall('Price'):
        CurrencyCode = price.find('CurrencyCode').text
        if CurrencyCode == 'GBP':
            gbp = price.find('PriceAmount').text
            print("GBP :: ", gbp)
            ws['C' + str(idx)] = gbp
        elif CurrencyCode == 'EUR':
            eur = price.find('PriceAmount').text
            print("EUR :: ", eur)
            ws['D' + str(idx)] = eur

wb.save("one.xml")
