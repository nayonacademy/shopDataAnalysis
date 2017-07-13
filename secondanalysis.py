from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
import xml.etree.ElementTree as ET
tree = ET.parse('shop.xml')
root = tree.getroot()
wb = Workbook()
ws = wb.active

for idx, child in enumerate(root.findall('Product')):
    DistinctiveTitle = child.find('DistinctiveTitle').text
    idx = idx +2
    col = 'A'+ str(idx)
    print("col :: ", col)
    ws[col] = DistinctiveTitle

    for id in child.findall('ProductIdentifier'):
        ProductIDType = id.find('ProductIDType').text
        if ProductIDType == '15':
            IDValue = id.find('IDValue').text
            ws['B'+ str(idx)] = IDValue

    SupplyDetail = child.find('SupplyDetail')
    for price in SupplyDetail.findall('Price'):
        CurrencyCode = price.find('CurrencyCode').text
        if CurrencyCode == 'GBP':
            gbp = price.find('PriceAmount').text
            ws['C' + str(idx)] = gbp
            gbp1 = float(gbp) * 1.13 #GBP to EUR google rate
            print("GBP :: ", gbp)
            ws['D' + str(idx)] = gbp1
            eur = SupplyDetail.findall('Price')[1].find('PriceAmount').text
            if gbp1 > float(eur):
                ft = Font(color=colors.RED)
                ws['F' + str(idx)] = "EUR"
                a1 = ws['F' + str(idx)]
                a1.font = ft
            else:
                ws['F' + str(idx)] = "GBP"
        elif CurrencyCode == 'EUR':
            eur = price.find('PriceAmount').text
            print("EUR :: ", eur)
            ws['E' + str(idx)] = eur

wb.save("two.xlsx")
